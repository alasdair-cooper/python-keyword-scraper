from datetime import datetime
import re
import os, os.path
from bs4.builder import ParserRejectedMarkup
import requests
import io
from mediawiki import MediaWiki
from sys import path
from string import ascii_uppercase
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from requests.exceptions import HTTPError
from requests.models import Response
from bs4 import BeautifulSoup

wikipedia = MediaWiki(user_agent="pyMediaWiki-ttnt59-data-cleaning-cw")

bbcUri = "https://www.bbc.co.uk/"
searchUri = "search?q="
pageUri = "&page="

def request(type: str, uri: str):
    try:
        response: Response
        if type.lower() == "get":
            response = requests.get(uri)
            if response.status_code == 200:
                return response.text
            else:
                raise HTTPError
    except HTTPError as http_err:
        print(f"HTTP error occurred: {http_err}")
    except Exception as err:
        print(f"Other error occurred: {err}")
    
def write_to_file(fileName: str, content: str):
    with io.open(fileName, 'w', encoding="utf-8") as f:
        f.write(content)

def read_from_file(fileName: str):
    with io.open(fileName, 'r', encoding="utf-8") as f:
        return f.read()

def search_results_raw(searchTerm: str, page: 1):
    return request("get", f"{bbcUri}{searchUri}{searchTerm}{pageUri}{page}")

def clean_results(contents: str):
    soup = BeautifulSoup(contents, "html.parser")
    tags = soup.find_all(["script", "style", "svg"])
    for tag in tags:
        tag.extract()
    return(soup)

def clean_article(contents: str):
    soup = BeautifulSoup(contents, "html.parser")
    tags = soup.find_all(["script", "style", "svg", "ul", "footer", "iframe"])
    for tag in tags:
        tag.extract()
    return(soup)

def get_individual_results(soup: BeautifulSoup):
    tags = soup.find_all(class_=re.compile("(PromoContent)"))
    return tags

def get_result_title(soup: BeautifulSoup):
    tags = soup.find_all("span")
    if len(tags) >= 1:
        return tags[0]
    else:
        return None

def get_result_link(soup: BeautifulSoup):
    tags = soup.find_all("a")
    if len(tags) >= 1:
        return tags[0]
    else:
        return None

def get_result_content(soup: BeautifulSoup):
    tags = soup.find_all("p")
    if len(tags) >= 2:
        return tags[1]
    else:
        return None

def parse_results(contents: str):
    results = []
    individualResults = get_individual_results(clean_results(contents))
    for result in individualResults:
        title = get_result_title(result)
        content = get_result_content(result)
        link = get_result_link(result)
        if(title != None and content != None):
            results.append(SearchResult(link.get("href"), content.string))
    return results

def store_article(path: str, link: str):
    write_to_file(path, request("get", f"{link}"))

def get_keywords():
    wb = load_workbook("keywords.xlsx")
    keywords = []
    index = 2
    value = wb["Sheet1"][f"A{index}"].value
    while value != None:
        keywords.append(value.lower())
        index += 1
        value = wb["Sheet1"][f"A{index}"].value
    return keywords

class SearchResult:
    def __init__(self, link, content):
        self.link = link
        self.content = content

## Problem 1

# BBC loops back to first page after requesting beyond the max page
# could filter out /av/ short articles+video but not enough content anyway
def problem1():
    keywords = get_keywords()
    for keyword in keywords:
        relevance = 1
        relevant = []
        page = 1
        raw = search_results_raw(keyword, page)
        soup = BeautifulSoup(raw, "html.parser")
        tags = soup.find_all(class_=re.compile("(PageButtonListItem)"))
        if len(tags) > 0:
            max = int(next(tags[-1].strings))
        else:
            max = 0
        while relevance > 0 and len(relevant) < 100 and page <= max:
            relevance = 0
            results = parse_results(search_results_raw(keyword, page).lower())
            for result in results:
                if "https://www.bbc.co.uk/news/" in result.link:
                    result.content = request("get", result.link)
                    if keyword.lower() in result.content:
                        relevant.append(result)
                        relevance += 1
            page += 1
        for result in relevant:
            split = str.split(result.link, "https://www.bbc.co.uk/news/")
            path = f"./pages/problem1/{keyword.replace(' ', '-')}.{split[1]}"
            write_to_file(path, result.content)

## Problem 2

def problem2():
    firstDirectory = ".\pages\problem1"
    secondDirectory = ".\pages\problem2"
    for filename in os.listdir(firstDirectory):
        soup = clean_article(read_from_file(os.path.join(firstDirectory, filename)))
        tags = soup.find_all("p")
        content = soup.find_all("h1")[0].string + "\n"
        for tag in tags:
            for string in tag.stripped_strings:
                if(string[-1] == '.'):
                    content += f"{string}\n"
                else:
                    content += f"{string} "
        write_to_file(os.path.join(secondDirectory, filename), content.lower())

## Problem 3

def problem3():
    directory = ".\pages\problem2"
    wb = load_workbook("keywords.xlsx")
    articleContents = {}
    for filename in os.listdir(directory):
        key = filename.split('.')[0].replace('-', ' ')
        if key in articleContents:
            articleContents[key].append(read_from_file(os.path.join(directory, filename)))
        else:
            articleContents[key] = [read_from_file(os.path.join(directory, filename))]
    
    row = 2
    rowKeyword = wb["Sheet1"][f"A{row}"].value.lower()
    keywordsChecked = []
    while rowKeyword != None:
        wikiPageContent = wikipedia.page(wikipedia.search(rowKeyword, results=1)[0]).content.lower()
        if rowKeyword in articleContents:
            for article in articleContents[rowKeyword]:
                column = 0
                columnKeyword = wb["Sheet1"][f"{ascii_uppercase[column]}1"].value.lower()
                while columnKeyword != None:
                    if columnKeyword != rowKeyword and not columnKeyword in keywordsChecked:
                        if wb["Sheet1"][f"{ascii_uppercase[column]}{row}"].value == None:
                            wb["Sheet1"][f"{ascii_uppercase[column]}{row}"].value = 1
                            wb["Sheet1"][f"{ascii_uppercase[row - 1]}{column + 1}"].value = 1
                        if columnKeyword in article:
                            wb["Sheet1"][f"{ascii_uppercase[column]}{row}"].value /= 2
                            wb["Sheet1"][f"{ascii_uppercase[row - 1]}{column + 1}"].value /= 2
                        if columnKeyword in wikiPageContent:
                            wb["Sheet1"][f"{ascii_uppercase[column]}{row}"].value /= 1.25
                            wb["Sheet1"][f"{ascii_uppercase[row - 1]}{column + 1}"].value /= 1.25
                    #else:
                        #wb["Sheet1"][f"{ascii_uppercase[column]}{row}"].value = ""
                        #wb["Sheet1"][f"{ascii_uppercase[row - 1]}{column + 1}"].value = ""
                    column += 1
                    columnKeyword = wb["Sheet1"][f"{ascii_uppercase[column]}1"].value
        row += 1
        rowKeyword = wb["Sheet1"][f"A{row}"].value
        keywordsChecked.append(rowKeyword)
    wb.save("./distance.xlsx")

def main():
    directory = "./pages/problem1"
    fileCount = len([name for name in os.listdir(directory) if os.path.isfile(os.path.join(directory, name))])
    #if fileCount == 0 or input("Fetch articles? [Y/N] ").lower() == "y":
    problem1()
    #if fileCount == 0 or input("Parse articles? [Y/N] ").lower() == "y":
    problem2()
    problem3()
    #write_to_file("./wikitest.txt", wikipedia.page(wikipedia.search("spyware", results=1)[0]).content.lower())

if __name__ == "__main__":
    main()